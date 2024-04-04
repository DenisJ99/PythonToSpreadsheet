import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def extract_data(text, target_pid):
    data = []
    lines = text.split('\n')
    current_pid = None
    current_tid = None
    current_name = None
    process_name = None
    ppid = None
    threceive_counts = {}
    thcondvar_counts = {}
    threply_counts = {}
    thsem_counts = {}
    thmutex_counts = {}
    thnanosleep_counts = {}
    

    
    for line in lines:
        pid_match = re.search(r'pid:(\d+)', line)
        if pid_match:
            current_pid = pid_match.group(1)
        
        tid_match = re.search(r'tid:(\d+)', line)
        if tid_match:
            current_tid = tid_match.group(1)
        
        name_match = re.search(r'name:(.+)', line)
        if name_match:
            current_name = name_match.group(1)
        
        ppid_match = re.search(r'ppid:(\d+)', line)
        if ppid_match:
            ppid = ppid_match.group(1)
        
        if current_pid == target_pid and current_name and current_name.startswith('./'):
            process_name = current_name
        
        if current_pid and current_tid and current_name:
            if current_pid == target_pid:
                data.append([current_name, current_tid])
            current_pid = None
            current_tid = None
            current_name = None
        
        if current_pid == target_pid and 'THRECEIVE' in line:
            if current_tid not in threceive_counts:
                threceive_counts[current_tid] = 0
            threceive_counts[current_tid] += 1
        
        if current_pid == target_pid and 'THCONDVAR' in line:
            if current_tid not in thcondvar_counts:
                thcondvar_counts[current_tid] = 0
            thcondvar_counts[current_tid] += 1
        
        if current_pid == target_pid and 'THREPLY' in line:
            if current_tid not in threply_counts:
                threply_counts[current_tid] = 0
            threply_counts[current_tid] += 1
        
        if current_pid == target_pid and 'THSEM' in line:
            if current_tid not in thsem_counts:
                thsem_counts[current_tid] = 0
            thsem_counts[current_tid] += 1

        if current_pid == target_pid and 'THMUTEX' in line:
            if current_tid not in thmutex_counts:
                thmutex_counts[current_tid] = 0
            thmutex_counts[current_tid] += 1

        if current_pid == target_pid and 'THNANOSLEEP' in line:
            if current_tid not in thnanosleep_counts:
                thnanosleep_counts[current_tid] = 0
            thnanosleep_counts[current_tid] += 1
    
    return data, process_name, ppid, threceive_counts, thcondvar_counts, threply_counts, thsem_counts, thmutex_counts, thnanosleep_counts

def set_cell_style(cell, font=None, alignment=None, fill=None, border=None):
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border

def write_to_xlsx(data, output_file, target_pid, process_name, ppid, threceive_counts, thcondvar_counts, threply_counts, thsem_counts, thmutex_counts, thnanosleep_counts): 
    workbook = Workbook()
    sheet = workbook.active
    
    # Define styles
    header_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal='center')
    fill_green = PatternFill(start_color='96d2b8', end_color='96d2b8', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Calculate the total THRECEIVE count
    total_threceive_count = sum(threceive_counts.values())

    # Calculate the total THCONDVAR count
    total_thcondvar_count = sum(thcondvar_counts.values())

    # Calculate the total THREPLY count
    total_threply_count = sum(threply_counts.values())

    # Calculate the total THSEM count
    total_thsem_count = sum(thsem_counts.values())

    # Calculate the total THMUTEX count
    total_thmutex_count = sum(thmutex_counts.values())

    # Calculate the total THNANOSLEEP count
    total_thnanosleep_count = sum(thnanosleep_counts.values())
    
    # Write headers for the first row
    headers_row1 = ["Process Name", "Running Time", "Process ID", "Running Time (Sum)",
                    "Running Time (Sum)", "Ready Time (Sum)", "Blocked Time", "Blocked Time (Sum)",
                    "Num Kernel Calls", "Num Messages", " ", "Running Time (msec)", "CPU Usage %",
                    "THRECEIVE",    "THCONDVAR",    "THREPLY",  "THSEM",    "THMUTEX", "THNANOSLEEP"]
    for col, header in enumerate(headers_row1, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)
    
    # Write headers for the third row
    headers_row3 = ["Thread Name", " ", "Thread ID", "Running Time (Sum)", " ", " ", " ", " ",
                    " ", " ", "Thread Owner ", "Running Time (msec)", "CPU Usage %",
                    "THRECEIVE"]
    for col, header in enumerate(headers_row3, start=1):
        cell = sheet.cell(row=3, column=col)
        cell.value = header
        set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)
    
    cell = sheet.cell(row=3, column=15) 
    cell.value = "THCONDVAR"
    set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)
    
    cell = sheet.cell(row=3, column=16)
    cell.value = "THREPLY"
    set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)

    cell = sheet.cell(row=3, column=17) 
    cell.value = "THSEM"
    set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)

    cell = sheet.cell(row=3, column=18) 
    cell.value = "THMUTEX"
    set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)

    cell = sheet.cell(row=3, column=19) 
    cell.value = "THNANOSLEEP"
    set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)
    
    # Write the process name in cell A2
    cell = sheet.cell(row=2, column=1)
    cell.value = process_name
    
    # Write the PPID in cell C2
    cell = sheet.cell(row=2, column=3)
    cell.value = ppid
    
    # Write the PID in cell C2
    cell = sheet.cell(row=2, column=3)
    cell.value = target_pid
    
    # Write data starting from the fourth row
    for row, item in enumerate(data, start=4):
        # Write the thread name in column A
        cell = sheet.cell(row=row, column=1)
        cell.value = item[0]
        
        # Write the thread ID in column C
        cell = sheet.cell(row=row, column=3)
        cell.value = item[1]
        
        # Write the THRECEIVE count for the corresponding thread ID in column N
        tid = item[1]
        if tid in threceive_counts:
            cell = sheet.cell(row=row, column=14)
            cell.value = threceive_counts[tid]
        
        # Write the THCONDVAR count for the corresponding thread ID in column 15
        if tid in thcondvar_counts:
            cell = sheet.cell(row=row, column=15) # Write to the new column for THCONDVAR
            cell.value = thcondvar_counts[tid]
        
        # Write the THREPLY count for the corresponding thread ID in column 16
        if tid in threply_counts:
            cell = sheet.cell(row=row, column=16) # Write to the new column for THREPLY
            cell.value = threply_counts[tid]

        # Write the THSEM count for the corresponding thread ID in column 17
        if tid in thsem_counts:
            cell = sheet.cell(row=row, column=17) # Write to the new column for THSEM
            cell.value = thsem_counts[tid]

        # Write the THMUTEX count for the corresponding thread ID in column 18
        if tid in thmutex_counts:
            cell = sheet.cell(row=row, column=18) # Write to the new column for THMUTEX
            cell.value = thmutex_counts[tid]

        # Write the THNANOSLEEP count for the corresponding thread ID in column 19
        if tid in thnanosleep_counts:
            cell = sheet.cell(row=row, column=19) # Write to the new column for THNANOSLEEP
            cell.value = thnanosleep_counts[tid]

        # Write the total THRECEIVE count to cell N2
        cell = sheet.cell(row=2, column=14)
        cell.value = total_threceive_count

         # Write the total THCONDVAR count to cell N2
        cell = sheet.cell(row=2, column=15)
        cell.value = total_thcondvar_count

         # Write the total THREPLY count to cell N2
        cell = sheet.cell(row=2, column=16)
        cell.value = total_threply_count

         # Write the total THSEM count to cell N2
        cell = sheet.cell(row=2, column=17)
        cell.value = total_thsem_count

         # Write the total THMUTEX count to cell N2
        cell = sheet.cell(row=2, column=18)
        cell.value = total_thmutex_count

        # Write the total THNANOSLEEP count to cell N2
        cell = sheet.cell(row=2, column=19)
        cell.value = total_thnanosleep_count
    
    adjust_column_widths(sheet)
    workbook.save(output_file)

# Read the text file
with open('17090_Cpu_Load_traceprintwide.txt', 'r') as file:
    text = file.read()

# Prompt the user to enter a pid
target_pid = input("Enter the pid to filter: ")

# Extract the data based on the specified pid
extracted_data, process_name, ppid, threceive_counts, thcondvar_counts, threply_counts, thsem_counts, thmutex_counts, thnanosleep_counts = extract_data(text, target_pid)

# Write the filtered data to an XLSX file
output_file = f'output_{target_pid}.xlsx'
write_to_xlsx(extracted_data, output_file, target_pid, process_name, ppid, threceive_counts, thcondvar_counts, threply_counts, thsem_counts, thmutex_counts, thnanosleep_counts)

print(f"Data filtered for pid {target_pid} has been written to {output_file}.")
